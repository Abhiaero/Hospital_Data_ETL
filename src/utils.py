# src/utils.py
import re
import json
import csv
from pathlib import Path
from typing import Tuple, Dict, List, Optional
import pandas as pd

KNOWN_HEADER_TOKENS = {
    "description", "code", "code|1", "code|1|type", "standard_charge|gross",
    "standard_charge|discounted_cash", "payer_name", "plan_name", "billing_class",
    "setting", "modifiers"
}

def normalize_colname(c: str) -> str:
    """Normalize column names to snake_case, replace '|' and spaces with underscores."""
    if c is None:
        return c
    c = str(c)
    c = c.strip().lower()
    c = c.replace("|", "_")
    c = re.sub(r"[^\w]+", "_", c)
    c = re.sub(r"_+", "_", c)
    c = c.strip("_")
    return c

def detect_header_line_csv(file_path: Path, n_preview: int = 25) -> Tuple[int, List[str]]:
    """
    Read the first n_preview lines and heuristically detect the header line index (0-based).
    Returns (header_line_index, preview_lines).
    """
    with open(file_path, "r", encoding="utf-8", errors="replace") as f:
        lines = [next(f) for _ in range(n_preview)]
    # check for a line that contains known header tokens
    for i, line in enumerate(lines):
        low = line.lower()
        # quick check: does it contain a token like 'description' or 'code|1'?
        for token in KNOWN_HEADER_TOKENS:
            if token in low:
                return i, lines
    # fallback: use the line with the largest number of delimiters (commas)
    max_commas = -1
    best_idx = 0
    for i, line in enumerate(lines):
        ccount = line.count(",")
        if ccount > max_commas:
            max_commas = ccount
            best_idx = i
    return best_idx, lines

def extract_metadata_from_lines(lines: List[str]) -> Dict[str, str]:
    """
    Given top-of-file lines (list of strings), attempt to extract metadata like hospital_name,
    last_updated_on, version, address, license_number. This is heuristic: many hospital files
    put a header row followed by a row with those values.
    """
    meta = {}
    # try to find line that contains 'hospital_name'
    for i, line in enumerate(lines):
        if "hospital_name" in line.lower().replace(" ", "") or "hospital name" in line.lower():
            # the next non-empty line often has the hospital values
            if i + 1 < len(lines):
                nxt = lines[i + 1].strip()
                # attempt to parse as CSV/quoted CSV
                try:
                    # use csv.reader to parse
                    data = next(csv.reader([nxt]))
                    header = next(csv.reader([line]))
                    # pair them
                    for k, v in zip(header, data):
                        k_n = normalize_colname(k)
                        meta[k_n] = v.strip()
                    return meta
                except Exception:
                    continue
    # fallback: try first two lines (many files put metadata in the first two)
    try:
        header = next(csv.reader([lines[0]]))
        data = next(csv.reader([lines[1]]))
        for k, v in zip(header, data):
            k_n = normalize_colname(k)
            meta[k_n] = v.strip()
        return meta
    except Exception:
        return meta

def read_generic(file_path: Path, nrows: Optional[int] = None) -> Tuple[pd.DataFrame, Dict]:
    """
    Robust reader for CSV, JSON, XLSX. Returns (df, metadata).
    For demo we use nrows to limit read size.
    """
    ext = file_path.suffix.lower()
    meta = {}
    if ext in {".csv", ".txt"}:
        header_idx, lines = detect_header_line_csv(file_path)
        meta = extract_metadata_from_lines(lines[: max(5, header_idx + 2)])
        # read with pandas using skiprows up to header_idx
        try:
            df = pd.read_csv(
                file_path,
                skiprows=header_idx,
                nrows=nrows,
                low_memory=False,
                dtype=str,
                engine="python",
            )
        except Exception:
            # fallback: read with default parameters and try to recover
            df = pd.read_csv(file_path, nrows=nrows, low_memory=False, dtype=str, engine="python", error_bad_lines=False)
        # normalize column names
        df.columns = [normalize_colname(c) for c in df.columns]
        return df, meta

    elif ext in {".json"}:
        # try multiple possible json layouts
        text = file_path.read_text(encoding="utf-8", errors="replace").strip()
        try:
            obj = json.loads(text)
            # if top-level has metadata fields
            if isinstance(obj, dict):
                # if there's a 'hospital_name' top-level key, extract it
                if "hospital_name" in obj:
                    meta = {normalize_colname(k): v for k, v in obj.items() if not isinstance(v, list)}
                # find a list of records inside the JSON
                for k, v in obj.items():
                    if isinstance(v, list):
                        df = pd.DataFrame(v)
                        df.columns = [normalize_colname(c) for c in df.columns]
                        if nrows:
                            df = df.head(nrows)
                        return df, meta
                # fallback: wrap dict into single-row DataFrame
                df = pd.json_normalize(obj)
                df.columns = [normalize_colname(c) for c in df.columns]
                if nrows:
                    df = df.head(nrows)
                return df, meta
            elif isinstance(obj, list):
                df = pd.DataFrame(obj)
                df.columns = [normalize_colname(c) for c in df.columns]
                if nrows:
                    df = df.head(nrows)
                return df, meta
        except json.JSONDecodeError:
            # maybe it's newline-delimited JSON
            try:
                df = pd.read_json(file_path, lines=True, nrows=nrows)
                df.columns = [normalize_colname(c) for c in df.columns]
                return df, meta
            except Exception:
                raise

    elif ext in {".xlsx", ".xls"}:
        # use pandas.read_excel but try to detect header row similarly
        # read first 10 rows as text with openpyxl to find header string
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, read_only=True)
            sheet = wb[wb.sheetnames[0]]
            preview = []
            header_row_idx = 0
            for i, row in enumerate(sheet.iter_rows(max_row=25, max_col=50, values_only=True), start=1):
                row_str = ",".join([str(c) if c is not None else "" for c in row])
                preview.append(row_str)
                if any(tok in row_str.lower() for tok in KNOWN_HEADER_TOKENS):
                    header_row_idx = i - 1  # pandas skiprows expects 0-based
                    break
            meta = extract_metadata_from_lines(preview[: max(5, header_row_idx + 2)])
            df = pd.read_excel(file_path, header=header_row_idx, nrows=nrows, engine="openpyxl", dtype=str)
            df.columns = [normalize_colname(c) for c in df.columns]
            return df, meta
        except Exception:
            # fallback to simple read
            df = pd.read_excel(file_path, nrows=nrows, engine="openpyxl", dtype=str)
            df.columns = [normalize_colname(c) for c in df.columns]
            return df, meta
    else:
        raise ValueError(f"Unsupported extension: {ext}")
